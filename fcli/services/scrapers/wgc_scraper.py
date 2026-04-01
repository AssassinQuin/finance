"""
World Gold Council (WGC) Supply/Demand Data Scraper.

Data source: https://www.gold.org/goldhub/data/demand-and-supply
Quarterly gold supply and demand statistics.

Excel Structure (GDT_Tables_Q425_CN.xlsx):
- Sheet "黄金供需" contains supply/demand data
- Row 5: Year/Quarter headers
- Rows 6-28: Data rows with labels in column B
- Columns 3-18: Annual data (2010-2025)
- Columns 24+: Quarterly data (2010 Q1 onwards)
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ...core.models.gold_supply_demand import GoldSupplyDemand
from ...infra.http_client import http_client
from ...utils.time_util import utcnow

logger = logging.getLogger(__name__)


# Excel Row Mapping (Row number -> Field name)
# Based on WGC Gold Demand Trends Excel format
ROW_MAPPING: dict[int, str] = {
    6: "total_supply",  # 总供应量
    7: "mine_production",  # 金矿产量
    8: "net_hedging",  # 生产商净套保
    10: "recycling",  # 回收金
    11: "total_demand",  # 总需求量
    12: "jewelry",  # 金饰制造
    15: "technology",  # 科技
    19: "total_investment",  # 投资
    20: "bars_coins",  # 金条和金币总需求量
    24: "etfs",  # 黄金ETF及类似产品
    25: "central_banks",  # 各央行和官方机构
    27: "otc_investment",  # 场外投资及其他
    28: "price_avg_usd",  # LBMA黄金价格
}


class WGCScraper:
    """WGC Gold Supply/Demand Data Scraper with URL discovery."""

    # Base URL pattern for WGC Excel files
    # Pattern observed: https://www.gold.org/sites/default/files/{YYYY}-{MM}/GDT_Tables_Q{Q}{YY}_CN.xlsx
    BASE_URL_TEMPLATE = (
        "https://www.gold.org/sites/default/files/{year}-{month:02d}/GDT_Tables_Q{quarter}{short_year}_CN.xlsx"
    )

    # Fallback URL pattern (alternative format)
    ALT_URL_TEMPLATE = "https://www.gold.org/sites/default/files/{year}-{month:02d}/GDT_Tables_Q{quarter}{year}_CN.xlsx"

    # Maximum number of quarters to try when discovering
    MAX_QUARTERS_TO_TRY = 12

    def __init__(self):
        self._last_successful_url: str | None = None

    async def close(self):
        """Close resources (no-op for WGCScraper, uses global http_client)"""
        pass

    def _build_url(self, year: int, quarter: int) -> str:
        """
        Build download URL for a given quarter.

        WGC typically releases Q1 data in late April, Q2 in late July,
        Q3 in late October, and Q4 in late January of the following year.

        Args:
            year: Year (e.g., 2024)
            quarter: Quarter (1-4)

        Returns:
            URL string for the Excel file
        """
        # Calculate expected release month for each quarter
        # Q1: released ~April, Q2: ~July, Q3: ~October, Q4: ~January (next year)
        release_months = {1: 4, 2: 7, 3: 10, 4: 11}  # Q4 often released in Nov/Dec

        month = release_months.get(quarter, 4)
        short_year = year % 100

        return self.BASE_URL_TEMPLATE.format(
            year=year,
            month=month,
            quarter=quarter,
            short_year=short_year,
        )

    def _build_alternative_url(self, year: int, quarter: int) -> str:
        """Build alternative URL format (full year in filename)."""
        release_months = {1: 4, 2: 7, 3: 10, 4: 11}
        month = release_months.get(quarter, 4)

        return self.ALT_URL_TEMPLATE.format(
            year=year,
            month=month,
            quarter=quarter,
        )

    def _recent_quarters(self, max_quarters: int = 12) -> list[tuple[int, int]]:
        """
        Generate recent (year, quarter) pairs in reverse chronological order.

        Args:
            max_quarters: Maximum number of quarters to generate

        Returns:
            List of (year, quarter) tuples, most recent first
        """
        now = datetime.now()
        current_year = now.year
        current_quarter = (now.month - 1) // 3 + 1

        quarters = []
        year, quarter = current_year, current_quarter

        for _ in range(max_quarters):
            quarters.append((year, quarter))
            quarter -= 1
            if quarter < 1:
                quarter = 4
                year -= 1

        return quarters

    async def _download_and_parse(self, url: str, year: int, quarter: int) -> list[GoldSupplyDemand] | None:
        """
        Download and parse Excel file from URL.

        Args:
            url: URL to download from
            year: Expected year (for logging)
            quarter: Expected quarter (for logging)

        Returns:
            List of GoldSupplyDemand records or None if failed
        """
        try:
            content = await http_client.get_binary(url)
            if content is None or len(content) < 1000:
                logger.debug(f"Invalid content from {url}")
                return None

            logger.info(f"Downloaded WGC Excel for {year} Q{quarter}: {len(content)} bytes")
            self._last_successful_url = url

            return self.parse_excel(content)

        except Exception as e:
            logger.debug(f"Failed to download/parse {url}: {e}")
            return None

    async def download_excel(self, year: int, quarter: int) -> bytes | None:
        """
        Download WGC Excel file for a specific quarter.

        Args:
            year: Year (e.g., 2024)
            quarter: Quarter (1-4)

        Returns:
            Excel file bytes or None if not available
        """
        # Try primary URL pattern
        url = self._build_url(year, quarter)
        content = await http_client.get_binary(url)
        if content and len(content) > 1000:
            logger.info(f"Downloaded WGC Excel for {year} Q{quarter}: {len(content)} bytes")
            return content

        # Try alternative URL pattern
        alt_url = self._build_alternative_url(year, quarter)
        content = await http_client.get_binary(alt_url)
        if content and len(content) > 1000:
            logger.info(f"Downloaded WGC Excel for {year} Q{quarter}: {len(content)} bytes")
            return content

        logger.debug(f"No valid Excel file found for {year} Q{quarter}")
        return None

    def parse_excel(self, excel_path: str | Path | bytes) -> list[GoldSupplyDemand]:
        """
        Parse WGC Excel file and extract quarterly supply/demand data.

        Args:
            excel_path: Path to Excel file or bytes content

        Returns:
            List of GoldSupplyDemand records
        """
        try:
            if isinstance(excel_path, bytes):
                workbook = load_workbook(BytesIO(excel_path), data_only=True)
            else:
                workbook = load_workbook(Path(excel_path), data_only=True)
        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            return []

        # Find the sheet with supply/demand data
        ws = None
        for sheet in workbook.worksheets:
            if sheet.title == "黄金供需":
                ws = sheet
                break

        if ws is None:
            logger.warning("Sheet '黄金供需' not found, trying first sheet")
            ws = workbook.active

        if ws is None:
            logger.error("No worksheet found in Excel file")
            return []

        # Validate expected rows exist
        self._validate_sheet_structure(ws)

        results: list[GoldSupplyDemand] = []

        # Find quarterly data columns (typically columns 24+)
        for col_idx in range(24, ws.max_column + 1):
            data = self._extract_quarter_data(ws, col_idx)
            if data:
                results.append(data)

        # Also try earlier columns in case structure changed
        if not results:
            for col_idx in range(3, ws.max_column + 1):
                data = self._extract_quarter_data(ws, col_idx)
                if data:
                    results.append(data)

        return results

    def _validate_sheet_structure(self, ws: Worksheet) -> None:
        """Validate that expected rows exist in the worksheet."""
        expected_labels = {
            6: "总供应量",
            11: "总需求量",
        }

        for row_idx, expected_label in expected_labels.items():
            cell_value = ws.cell(row=row_idx, column=2).value
            if cell_value is None or expected_label not in str(cell_value):
                logger.warning(f"Expected '{expected_label}' at row {row_idx}, got '{cell_value}'")

    def _extract_quarter_data(self, ws: Worksheet, col_idx: int) -> GoldSupplyDemand | None:
        """Extract data for a specific quarter from a column."""
        # Check if column has a year/quarter header (e.g., "2024 Q1")
        cell_value = ws.cell(row=5, column=col_idx).value
        if cell_value is None:
            return None

        # Parse quarter from cell value like "2024 Q1"
        header_str = str(cell_value).strip()

        # Try to parse "YYYY QN" format
        match = re.match(r"(\d{4})\s*Q(\d)", header_str, re.IGNORECASE)
        if not match:
            return None

        try:
            year = int(match.group(1))
            quarter = int(match.group(2))
        except (ValueError, TypeError):
            return None

        # Validate quarter
        if quarter < 1 or quarter > 4:
            return None

        # Read all values from the column
        values: dict[str, float] = {}

        for row_idx, field_name in ROW_MAPPING.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                try:
                    values[field_name] = float(cell_value)
                except (ValueError, TypeError):
                    pass

        # Skip if no meaningful data
        if not values or values.get("total_supply", 0) == 0:
            return None

        # Calculate balance
        balance = values.get("total_supply", 0) - values.get("total_demand", 0)

        return GoldSupplyDemand(
            year=year,
            quarter=quarter,
            period=f"{year} Q{quarter}",
            mine_production=values.get("mine_production", 0),
            recycling=values.get("recycling", 0),
            net_hedging=values.get("net_hedging", 0),
            total_supply=values.get("total_supply", 0),
            jewelry=values.get("jewelry", 0),
            technology=values.get("technology", 0),
            total_investment=values.get("total_investment", 0),
            bars_coins=values.get("bars_coins", 0),
            etfs=values.get("etfs", 0),
            otc_investment=values.get("otc_investment", 0),
            central_banks=values.get("central_banks", 0),
            total_demand=values.get("total_demand", 0),
            supply_demand_balance=balance,
            price_avg_usd=values.get("price_avg_usd"),
            data_source="WGC",
            fetch_time=utcnow(),
        )

    async def fetch_latest(self) -> list[GoldSupplyDemand]:
        """
        Fetch latest available quarterly data using URL discovery.

        Tries recent quarters in reverse chronological order until
        a valid Excel file is found.

        Returns:
            List of GoldSupplyDemand records
        """
        quarters = self._recent_quarters(self.MAX_QUARTERS_TO_TRY)

        for year, quarter in quarters:
            # Try primary URL pattern
            url = self._build_url(year, quarter)
            data = await self._download_and_parse(url, year, quarter)
            if data:
                return data

            # Try alternative URL pattern
            alt_url = self._build_alternative_url(year, quarter)
            data = await self._download_and_parse(alt_url, year, quarter)
            if data:
                return data

        logger.warning("No valid WGC Excel file found in recent quarters")
        return []

    def fetch_from_local(self, file_path: str | Path) -> list[GoldSupplyDemand]:
        """
        Parse a local WGC Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of GoldSupplyDemand records
        """
        return self.parse_excel(file_path)
